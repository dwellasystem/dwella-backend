from rest_framework import generics, filters, status
from rest_framework.permissions import AllowAny, IsAuthenticated
from .models import MonthlyBill
from units.models import AssignedUnit, Unit
from rest_framework.views import APIView
from django.db.models import Sum, Count
from django.db import models
from django.http import HttpResponse
from .serializers import MonthlyBillSerializer
from decimal import Decimal
from rest_framework.response import Response
import calendar
from django.db.models import Avg
from datetime import datetime, date
from django.db.models.functions import ExtractMonth, ExtractYear
from django_filters.rest_framework import DjangoFilterBackend
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl import Workbook
import io


class MonthlyBillListCreateView(generics.ListCreateAPIView):
    serializer_class = MonthlyBillSerializer
    permission_classes = [AllowAny]   # ✅ anyone can access
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['user', 'due_date', 'payment_status', 'due_status', 'sms_sent']
    search_fields = [
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__middle_name",
        "unit__unit_name",
        "user__email",
        "due_date", 
        "payment_status",
        "due_status"]
    ordering_fields = ['due_date', 'amount_due', 'unit__building']
    ordering = ['-due_date']  # default order

    def get_queryset(self):
        # ✅ show all bills, newest first
        return MonthlyBill.objects.all().order_by("-due_date")

    def perform_create(self, serializer):
        user_id = self.request.data.get("user_id")

        # 🧩 Case 1: Admin or employee manually creating a bill
        if self.request.user.is_authenticated and self.request.user.role in ["admin", "employee"]:
            if user_id:
                serializer.save()  # user assigned via user_id
            else:
                raise ValueError("You must specify which resident the bill is for.")

        # 🧩 Case 2: Celery or backend logic auto-generating (not authenticated)
        elif user_id:
            serializer.save()  # accept system/manual POSTs that explicitly set user_id

        # 🚫 Case 3: Resident trying to create a bill
        elif self.request.user.is_authenticated and self.request.user.role == "resident":
            raise PermissionError("Residents are not allowed to create bills.")

        # 🚫 Case 4: Anonymous user with no user_id
        else:
            raise ValueError("A user must be specified to create a bill.")


class MonthlyBillListView(generics.ListAPIView):
    queryset = MonthlyBill.objects.all().order_by('-due_date')
    serializer_class = MonthlyBillSerializer
    permission_classes = [IsAuthenticated]  # ✅ only authenticated users
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields = ['user', 'due_date', 'payment_status', 'due_status']
    search_fields = [
        "user__username",
        "user__first_name",
        "user__last_name",
        "user__middle_name",
        "user__email",
        "due_date", 
        "payment_status",
        "due_status"]
    ordering_fields = ['due_date', 'amount_due']
    ordering = ['-due_date']  # default order
    pagination_class = None  # ✅ no pagination


class MonthlyBillDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = MonthlyBillSerializer
    permission_classes = [AllowAny]   # ✅ anyone can access
    filter_backends = [filters.SearchFilter]
    search_fields = ["due_date", "payment_status", "due_status"]

    def get_queryset(self):
        # ✅ show all bills
        return MonthlyBill.objects.all()


# class MonthlyBillStatsView(APIView):
#     permission_classes = [AllowAny]

#     def get(self, request):
#         # ✅ Overdue bills only
#         overdue_bills = MonthlyBill.objects.filter(due_status="overdue")

#         total_overdue_count = overdue_bills.count()
#         total_overdue_amount = overdue_bills.aggregate(Sum("amount_due"))["amount_due__sum"] or 0

#         # ✅ Paid and pending counts (for small dashboard charts)
#         paid_count = MonthlyBill.objects.filter(payment_status="paid").count()
#         pending_count = MonthlyBill.objects.filter(payment_status="pending").count()

#         data = {
#             "total_overdue_count": total_overdue_count,
#             "total_overdue_amount": total_overdue_amount,
#             "paid_count": paid_count,
#             "pending_count": pending_count,
#         }

#         return Response(data)

class UnitStatusSummaryView(APIView):
    def get(self, request):
        statuses = dict(AssignedUnit.UnitStatus.choices)

        # Count each type
        data = {
            'owner_occupied': AssignedUnit.objects.filter(unit_status='owner_occupied').count(),
            'rented_short_term': AssignedUnit.objects.filter(unit_status='rented_short_term').count(),
            'air_bnb': AssignedUnit.objects.filter(unit_status='air_bnb').count(),
        }

        return Response(data)

class MonthlyBillSummaryView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        """
        Returns the total due and total collected payments for the current month.
        Example:
        {
            "month": "October",
            "year": 2025,
            "totalDue": 6000.00,
            "totalCollectedPayment": 4500.00,
            "totalOverDue": 2,
            "totalPending": 4
        }
        """
        # ✅ Get current month and year
        today = date.today()
        current_month = today.month
        current_year = today.year

        # ✅ Filter all bills for this month & year
        bills_this_month = MonthlyBill.objects.filter(
            due_date__year=current_year,
            due_date__month=current_month
        )

        # ✅ Calculate totals
        total_due = bills_this_month.filter(payment_status='pending').aggregate(total=Sum('amount_due'))['total'] or 0
        total_collected = bills_this_month.filter(payment_status='paid').aggregate(total=Sum('amount_due'))['total'] or 0
        total_overdue = bills_this_month.filter(due_status='overdue').count()
        all_overdue_bills = MonthlyBill.objects.filter(due_status='overdue').count()
        total_pending = bills_this_month.filter(payment_status='pending').aggregate(total=Sum('amount_due'))['total']  or 0

        # ✅ Prepare response
        data = {
            "month": calendar.month_name[current_month],
            "year": current_year,
            "totalDue": round(total_due, 2),
            "totalCollectedPayment": round(total_collected, 2),
            "totalOverDue": total_overdue,
            "totalPending": total_pending,
            "all_overdue_bills": all_overdue_bills
        }

        return Response(data, status=status.HTTP_200_OK)
    

class MonthlyBillStatsView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        """
        Returns monthly billing stats grouped by year and month.
        Example:
        [
          {"year": 2025, "month": "August", "paid_count": 3, "pending_count": 2, "overdue_count": 1},
          {"year": 2025, "month": "September", "paid_count": 5, "pending_count": 4, "overdue_count": 0},
          {"year": 2025, "month": "October", "paid_count": 2, "pending_count": 2, "overdue_count": 0}
        ]
        """

        # ✅ Annotate each bill with year and month
        bills = (
            MonthlyBill.objects
            .annotate(year=ExtractYear('due_date'), month=ExtractMonth('due_date'))
            .values('year', 'month')
            .annotate(
                paid_count=Count('id', filter=models.Q(payment_status='paid')),
                pending_count=Count('id', filter=models.Q(payment_status='pending')),
                overdue_count=Count('id', filter=models.Q(due_status='overdue')),
                total_paid=Sum('amount_due', filter=models.Q(payment_status='paid')),
                total_pending=Sum('amount_due', filter=models.Q(payment_status='pending')),
            )
            .order_by('year', 'month')
        )

        # ✅ Convert month numbers to month names
        formatted_data = [
            {
                "year": item["year"],
                "month": calendar.month_name[item["month"]],
                "paid_count": item["paid_count"],
                "pending_count": item["pending_count"],
                "overdue_count": item["overdue_count"],
                "total_paid": item["total_paid"] or 0,
                "total_pending": item["total_pending"] or 0,
            }
            for item in bills
        ]

        return Response(formatted_data)
    

class OverdueUserSummaryView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        """
        Returns a summary of all users with overdue bills:
        [
          {
            "user": "Eimann Calderon",
            "unit": ["Unit 201", "Unit 202"],
            "totalAmountDue": 4200.00,
            "monthsDue": ["September", "October"]
          },
          {
            "user": "Daniel Paddilla",
            "unit": ["Unit 101", "Unit 102"],
            "totalAmountDue": 60000.00,
            "monthsDue": ["September", "October"]
          }
        ]
        """

        # ✅ Restrict to admin/employee only
        if request.user.role not in ["admin", "employee"]:
            return Response(
                {"detail": "You are not authorized to view this data."},
                status=status.HTTP_403_FORBIDDEN
            )

        # ✅ Get all overdue bills
        overdue_bills = MonthlyBill.objects.filter(due_status="overdue")

        if not overdue_bills.exists():
            return Response([], status=status.HTTP_200_OK)

        # ✅ Group by user
        user_summaries = []

        # Get all unique users who have overdue bills
        users = overdue_bills.values_list("user", flat=True).distinct()

        for user_id in users:
            user_bills = overdue_bills.filter(user_id=user_id)
            user = user_bills.first().user

            # ✅ Get user full name or fallback to username
            user_name = f"{user.first_name} {user.last_name}".strip() or user.username

            # ✅ Get all unique units for this user
            units = list(user_bills.values_list("unit__unit_name", flat=True).distinct())

            # ✅ Calculate total overdue amount
            total_amount_due = (
                user_bills.aggregate(total=Sum("amount_due"))["total"] or 0
            )

            # ✅ Get months of overdue bills
            months = list(
                user_bills.annotate(month=ExtractMonth("due_date"))
                .values_list("month", flat=True)
                .distinct()
            )
            months_due = [calendar.month_name[m] for m in months]

            user_summaries.append({
                "user": user_name,
                "unit": units,
                "totalAmountDue": round(total_amount_due, 2),
                "monthsDue": months_due,
            })

        return Response(user_summaries, status=status.HTTP_200_OK)
    

class UserYearlyBillSummaryView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, user_id=None):
        """
        Returns detailed yearly bill summary for a specific user.
        """
        if not user_id:
            return Response(
                {"error": "User ID is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        today = date.today()
        current_year = today.year

        try:
            # Filter bills for this year and user
            bills_this_year = MonthlyBill.objects.filter(
                due_date__year=current_year,
                user_id=user_id
            )

            # ✅ Main totals - using Q objects for better filtering
            paid_bills = bills_this_year.filter(payment_status='paid')
            pending_bills = bills_this_year.filter(payment_status='pending')
            
            total_paid = paid_bills.aggregate(total=Sum('amount_due'))['total'] or 0
            total_unpaid = pending_bills.aggregate(total=Sum('amount_due'))['total'] or 0
            
            # ✅ Counts
            paid_bills_count = paid_bills.count()
            unpaid_bills_count = pending_bills.count()
            total_bills = paid_bills_count + unpaid_bills_count
            
            # ✅ Success rate
            payment_success_rate = round((paid_bills_count / total_bills * 100), 2) if total_bills > 0 else 0

            # ✅ Monthly breakdown - improved to handle months with no data
            monthly_data = []
            for month in range(1, 13):
                month_bills = bills_this_year.filter(due_date__month=month)
                month_paid = month_bills.filter(payment_status='paid').aggregate(total=Sum('amount_due'))['total'] or 0
                month_unpaid = month_bills.filter(payment_status='pending').aggregate(total=Sum('amount_due'))['total'] or 0
                month_total = month_paid + month_unpaid
                
                monthly_data.append({
                    "month": calendar.month_name[month],
                    "month_number": month,
                    "paid": round((month_paid), 2),
                    "unpaid": round((month_unpaid), 2),
                    "total": round((month_total), 2),
                    "bills_count": month_bills.count()
                })

            # ✅ Additional insights
            # Get unique units for this user in the current year
            unique_units = bills_this_year.values('unit__unit_name', 'unit__building').distinct()
            
            # Get overdue bills count for the current year
            overdue_bills_count = bills_this_year.filter(due_status='overdue').count()

            data = {
                "year": current_year,
                "user_id": user_id,
                "summary": {
                    "total_paid": round((total_paid), 2),
                    "total_unpaid": round((total_unpaid), 2),
                    "paid_bills_count": paid_bills_count,
                    "unpaid_bills_count": unpaid_bills_count,
                    "payment_success_rate": payment_success_rate,
                    "total_bills_count": total_bills,
                    "overdue_bills_count": overdue_bills_count,
                    "unique_units_count": len(unique_units),
                    "unique_units": list(unique_units)
                },
                "monthly_breakdown": monthly_data
            }

            return Response(data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"An error occurred: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        

class UserYearlyPaymentBreakdownView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, user_id=None):
        """
        Returns detailed payment breakdown for a specific user for the current year.
        Shows percentages of total amount, rent, and additional charges.
        If move-in date is in July, first bill collection is in August.
        """
        if not user_id:
            return Response(
                {"error": "User ID is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        # Get year from query parameter or use current year
        year_param = request.GET.get('year')
        if year_param:
            try:
                current_year = int(year_param)
            except ValueError:
                return Response(
                    {"error": "Invalid year format"}, 
                    status=status.HTTP_400_BAD_REQUEST
                )
        else:
            current_year = date.today().year

        try:
            # Get all bills for the specified year
            bills_this_year = MonthlyBill.objects.filter(
                due_date__year=current_year,
                user_id=user_id
            )

            # Get assigned units for this user to calculate additional charges
            assigned_units = AssignedUnit.objects.filter(
                assigned_by_id=user_id,
                deleted_at__isnull=True
            ).select_related('unit_id')

            # ✅ Calculate total amounts from bills - Convert to float for calculations
            total_paid_result = bills_this_year.filter(payment_status='paid').aggregate(
                total=Sum('amount_due')
            )['total'] or Decimal('0.00')
            total_unpaid_result = bills_this_year.filter(payment_status='pending').aggregate(
                total=Sum('amount_due')
            )['total'] or Decimal('0.00')
            
            total_paid = float(total_paid_result)
            total_unpaid = float(total_unpaid_result)
            total_amount = total_paid + total_unpaid

            # ✅ Calculate ACTUAL base rent and additional charges from bills
            base_rent_total = 0.0
            additional_charges_total = 0.0
            security_total = 0.0
            amenities_total = 0.0
            maintenance_total = 0.0

            # Configuration for charge amounts
            CHARGE_AMOUNTS = {
                'security': 2000.0,
                'amenities': 2500.0,
                'maintenance': 1500.0,
            }

            # Calculate what the charges should be based on assigned units and move-in date
            # WITH ONE-MONTH DELAY: If move-in is July, billing starts in August
            for assigned in assigned_units:
                if assigned.unit_id and assigned.unit_id.rent_amount:
                    # Get move-in date (use move_in_date or created_at as fallback)
                    move_in_date = assigned.move_in_date or assigned.created_at.date()
                    
                    # Only calculate for units that were assigned in the current year or earlier
                    if move_in_date.year <= current_year:
                        # Calculate starting month for billing (move-in month + 1)
                        if move_in_date.year == current_year:
                            # Started this year - billing starts from next month
                            start_month = move_in_date.month + 1
                            
                            # If move-in was in December, billing starts next year
                            if start_month > 12:
                                # No bills this year
                                months_remaining = 0
                            else:
                                months_remaining = 13 - start_month
                        else:
                            # Started in previous year - full year expected
                            months_remaining = 12
                        
                        # Base rent for the remaining months - Convert to float
                        monthly_rent = float(assigned.unit_id.rent_amount)
                        base_rent_total += monthly_rent * months_remaining
                        
                        # Additional charges for the remaining months
                        if assigned.security:
                            security_total += CHARGE_AMOUNTS['security'] * months_remaining
                        if assigned.amenities:
                            amenities_total += CHARGE_AMOUNTS['amenities'] * months_remaining
                        if assigned.maintenance:
                            maintenance_total += CHARGE_AMOUNTS['maintenance'] * months_remaining

            # Total additional charges
            additional_charges_total = security_total + amenities_total + maintenance_total
            
            # Expected total for the year (base rent + additional charges)
            expected_yearly_total = base_rent_total + additional_charges_total

            # ✅ Calculate ACTUAL breakdown from bills (this is the key fix)
            actual_base_rent_total = 0.0
            actual_additional_charges_total = 0.0
            
            for assigned in assigned_units:
                if assigned.unit_id and assigned.unit_id.rent_amount:
                    move_in_date = assigned.move_in_date or assigned.created_at.date()
                    
                    # Only consider units that were assigned in the current year or earlier
                    if move_in_date.year <= current_year:
                        # Get bills for this specific unit
                        unit_bills = bills_this_year.filter(unit_id=assigned.unit_id.id)
                        
                        if unit_bills.exists():
                            # Calculate expected monthly total for this unit
                            monthly_rent = float(assigned.unit_id.rent_amount)
                            monthly_additional = 0.0
                            
                            if assigned.security:
                                monthly_additional += CHARGE_AMOUNTS['security']
                            if assigned.amenities:
                                monthly_additional += CHARGE_AMOUNTS['amenities']
                            if assigned.maintenance:
                                monthly_additional += CHARGE_AMOUNTS['maintenance']
                            
                            monthly_total_expected = monthly_rent + monthly_additional
                            
                            # Calculate the ratio of base rent to total expected
                            if monthly_total_expected > 0:
                                base_rent_ratio = monthly_rent / monthly_total_expected
                                
                                # Apply this ratio to actual bill amounts
                                unit_total_paid = float(unit_bills.filter(payment_status='paid').aggregate(
                                    total=Sum('amount_due')
                                )['total'] or Decimal('0.00'))
                                
                                unit_total_unpaid = float(unit_bills.filter(payment_status='pending').aggregate(
                                    total=Sum('amount_due')
                                )['total'] or Decimal('0.00'))
                                
                                unit_total = unit_total_paid + unit_total_unpaid
                                
                                actual_base_rent_total += unit_total * base_rent_ratio
                                actual_additional_charges_total += unit_total * (1 - base_rent_ratio)
            
            # If we couldn't calculate actual breakdown from bills, use the expected values
            if actual_base_rent_total == 0 and actual_additional_charges_total == 0:
                actual_base_rent_total = base_rent_total
                actual_additional_charges_total = additional_charges_total

            # ✅ Calculate percentages based on ACTUAL amounts
            actual_expected_total = actual_base_rent_total + actual_additional_charges_total
            
            if actual_expected_total > 0:
                base_rent_percentage = round((actual_base_rent_total / actual_expected_total) * 100, 2)
                additional_charges_percentage = round((actual_additional_charges_total / actual_expected_total) * 100, 2)
                
                # Calculate individual additional charge percentages
                security_percentage = round((security_total / actual_expected_total) * 100, 2) if security_total > 0 else 0
                amenities_percentage = round((amenities_total / actual_expected_total) * 100, 2) if amenities_total > 0 else 0
                maintenance_percentage = round((maintenance_total / actual_expected_total) * 100, 2) if maintenance_total > 0 else 0
                
                # Actual paid vs expected
                paid_percentage_of_expected = round((total_paid / actual_expected_total) * 100, 2) if actual_expected_total > 0 else 0
                unpaid_percentage_of_expected = round((total_unpaid / actual_expected_total) * 100, 2) if actual_expected_total > 0 else 0
            else:
                base_rent_percentage = 0
                additional_charges_percentage = 0
                security_percentage = 0
                amenities_percentage = 0
                maintenance_percentage = 0
                paid_percentage_of_expected = 0
                unpaid_percentage_of_expected = 0

            # ✅ Monthly breakdown with percentages - ADJUSTED FOR ONE-MONTH DELAY
            monthly_breakdown = []
            for month in range(1, 13):
                month_bills = bills_this_year.filter(due_date__month=month)
                
                # Convert all amounts to float to avoid Decimal/float operations
                month_paid_result = month_bills.filter(payment_status='paid').aggregate(total=Sum('amount_due'))['total'] or Decimal('0.00')
                month_unpaid_result = month_bills.filter(payment_status='pending').aggregate(total=Sum('amount_due'))['total'] or Decimal('0.00')
                
                month_paid = float(month_paid_result)
                month_unpaid = float(month_unpaid_result)
                month_total = month_paid + month_unpaid
                
                # Calculate expected amount for this month with ONE-MONTH DELAY
                monthly_expected = 0.0
                
                # Method 1: Use actual bill amounts if bills exist for this month
                if month_bills.exists():
                    # If bills exist, use the sum of their amount_due as expected
                    month_expected_result = month_bills.aggregate(total=Sum('amount_due'))['total'] or Decimal('0.00')
                    monthly_expected = float(month_expected_result)
                else:
                    # Method 2: Calculate based on assigned units with ONE-MONTH DELAY
                    for assigned in assigned_units:
                        move_in_date = assigned.move_in_date or assigned.created_at.date()
                        
                        # Check if this month should be expected with one-month delay
                        # If move-in is July, billing starts in August (month >= move_in_month + 1)
                        if move_in_date.year < current_year:
                            # Previous year move-in: all months are expected
                            if assigned.unit_id and assigned.unit_id.rent_amount:
                                monthly_rent = float(assigned.unit_id.rent_amount)
                                monthly_expected += monthly_rent
                                
                                # Use the same charge amounts for monthly calculation
                                if assigned.security:
                                    monthly_expected += CHARGE_AMOUNTS['security']
                                if assigned.amenities:
                                    monthly_expected += CHARGE_AMOUNTS['amenities']
                                if assigned.maintenance:
                                    monthly_expected += CHARGE_AMOUNTS['maintenance']
                        elif move_in_date.year == current_year:
                            # Current year move-in: billing starts from next month
                            if month >= (move_in_date.month + 1):
                                if assigned.unit_id and assigned.unit_id.rent_amount:
                                    monthly_rent = float(assigned.unit_id.rent_amount)
                                    monthly_expected += monthly_rent
                                    
                                    # Use the same charge amounts for monthly calculation
                                    if assigned.security:
                                        monthly_expected += CHARGE_AMOUNTS['security']
                                    if assigned.amenities:
                                        monthly_expected += CHARGE_AMOUNTS['amenities']
                                    if assigned.maintenance:
                                        monthly_expected += CHARGE_AMOUNTS['maintenance']
                
                # Calculate percentage
                if monthly_expected > 0:
                    month_percentage_of_expected = round((month_total / monthly_expected) * 100, 2)
                else:
                    month_percentage_of_expected = 0

                monthly_breakdown.append({
                    "month": calendar.month_name[month],
                    "month_number": month,
                    "paid": round(month_paid, 2),
                    "unpaid": round(month_unpaid, 2),
                    "total": round(month_total, 2),
                    "expected_amount": round(monthly_expected, 2),
                    "percentage_of_expected": month_percentage_of_expected,
                    "bills_count": month_bills.count(),
                    "billing_note": "Billing starts one month after move-in" if monthly_expected > 0 else ""
                })

            # ✅ Unit-wise breakdown (considering move-in date with ONE-MONTH DELAY)
            unit_breakdown = []
            unique_units = bills_this_year.values('unit_id', 'unit__unit_name', 'unit__building').distinct()
            
            for unit_data in unique_units:
                unit_bills = bills_this_year.filter(unit_id=unit_data['unit_id'])
                
                # Convert all amounts to float
                unit_paid_result = unit_bills.filter(payment_status='paid').aggregate(total=Sum('amount_due'))['total'] or Decimal('0.00')
                unit_unpaid_result = unit_bills.filter(payment_status='pending').aggregate(total=Sum('amount_due'))['total'] or Decimal('0.00')
                
                unit_paid = float(unit_paid_result)
                unit_unpaid = float(unit_unpaid_result)
                unit_total = unit_paid + unit_unpaid
                
                # Find assigned unit to get additional charges and move-in date
                assigned_unit = assigned_units.filter(unit_id=unit_data['unit_id']).first()
                unit_expected_yearly = 0.0
                
                # Calculate expected amount based on ACTUAL bills for this unit
                if unit_bills.exists():
                    # Use actual bill amounts if bills exist
                    unit_expected_result = unit_bills.aggregate(total=Sum('amount_due'))['total'] or Decimal('0.00')
                    unit_expected_yearly = float(unit_expected_result)
                elif assigned_unit and assigned_unit.unit_id:
                    # Fall back to calculation based on assigned unit WITH ONE-MONTH DELAY
                    move_in_date = assigned_unit.move_in_date or assigned_unit.created_at.date()
                    
                    # Calculate remaining months in the year from move-in date + 1 month
                    if move_in_date.year == current_year:
                        start_month = move_in_date.month + 1  # One month delay
                        
                        # If move-in was in December, billing starts next year
                        if start_month > 12:
                            months_remaining = 0
                        else:
                            months_remaining = 13 - start_month
                    elif move_in_date.year < current_year:
                        months_remaining = 12
                    else:
                        months_remaining = 0
                    
                    unit_monthly_rent = float(assigned_unit.unit_id.rent_amount)
                    unit_expected_yearly = unit_monthly_rent * months_remaining
                    
                    # Use the same configurable amounts
                    if assigned_unit.security:
                        unit_expected_yearly += CHARGE_AMOUNTS['security'] * months_remaining
                    if assigned_unit.amenities:
                        unit_expected_yearly += CHARGE_AMOUNTS['amenities'] * months_remaining
                    if assigned_unit.maintenance:
                        unit_expected_yearly += CHARGE_AMOUNTS['maintenance'] * months_remaining
                
                unit_percentage_of_expected = round((unit_total / unit_expected_yearly) * 100, 2) if unit_expected_yearly > 0 else 0

                unit_breakdown.append({
                    "unit_id": unit_data['unit_id'],
                    "unit_name": unit_data['unit__unit_name'],
                    "building": unit_data['unit__building'],
                    "total_paid": round(unit_paid, 2),
                    "total_unpaid": round(unit_unpaid, 2),
                    "total_amount": round(unit_total, 2),
                    "expected_yearly_amount": round(unit_expected_yearly, 2),
                    "percentage_of_expected": unit_percentage_of_expected,
                    "bills_count": unit_bills.count(),
                    "billing_note": "First bill due one month after move-in" if unit_expected_yearly > 0 else ""
                })

            # ✅ Recalculate summary to ensure consistency
            recalculated_expected_yearly_total = sum(month['expected_amount'] for month in monthly_breakdown)
            
            if recalculated_expected_yearly_total > 0:
                recalculated_paid_percentage = round((total_paid / recalculated_expected_yearly_total) * 100, 2)
                recalculated_unpaid_percentage = round((total_unpaid / recalculated_expected_yearly_total) * 100, 2)
            else:
                recalculated_paid_percentage = 0
                recalculated_unpaid_percentage = 0

            data = {
                "year": current_year,
                "user_id": user_id,
                "summary": {
                    "total_paid": round(total_paid, 2),
                    "total_unpaid": round(total_unpaid, 2),
                    "total_amount": round(total_amount, 2),
                    "expected_yearly_total": round(recalculated_expected_yearly_total, 2),
                    "paid_percentage_of_expected": recalculated_paid_percentage,
                    "unpaid_percentage_of_expected": recalculated_unpaid_percentage,
                    "completion_rate": recalculated_paid_percentage,
                    "billing_policy": "First billing occurs one month after move-in date"
                },
                "breakdown_percentages": {
                    "base_rent": {
                        "amount": round(actual_base_rent_total, 2),
                        "percentage": base_rent_percentage
                    },
                    "additional_charges": {
                        "amount": round(actual_additional_charges_total, 2),
                        "percentage": additional_charges_percentage,
                        "details": {
                            "security": {
                                "amount": round(security_total, 2),
                                "percentage": security_percentage
                            },
                            "amenities": {
                                "amount": round(amenities_total, 2),
                                "percentage": amenities_percentage
                            },
                            "maintenance": {
                                "amount": round(maintenance_total, 2),
                                "percentage": maintenance_percentage
                            }
                        }
                    }
                },
                "monthly_breakdown": monthly_breakdown,
                "unit_breakdown": unit_breakdown,
                "debug_info": {
                    "charge_amounts_used": CHARGE_AMOUNTS,
                    "assigned_units_count": assigned_units.count(),
                    "bills_count": bills_this_year.count(),
                    "original_expected_total": round(expected_yearly_total, 2),
                    "recalculated_expected_total": round(recalculated_expected_yearly_total, 2),
                    "actual_base_rent_total": round(actual_base_rent_total, 2),
                    "actual_additional_charges_total": round(actual_additional_charges_total, 2),
                    "calculation_method": "proportional_allocation",
                    "billing_delay_months": 1
                }
            }

            return Response(data, status=status.HTTP_200_OK)

        except Exception as e:
            return Response(
                {"error": f"An error occurred: {str(e)}"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )   

from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status
from django.db.models import Sum, Count, Avg, Q
from django.db.models.functions import ExtractMonth
from datetime import date
import calendar

class FinancialReportsView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request, user_id=None):
        """
        Generate financial reports (monthly or yearly) for a specific user.
        Query Parameters:
        - period: 'monthly' or 'yearly' (default: 'monthly')
        - year: specific year (default: current year)
        - month: specific month (for monthly reports, default: current month)
        - breakdown: 'detailed' or 'summary' (default: 'detailed')
        """
        # Get user_id from URL path or query params
        user_id = user_id or request.GET.get('user_id')
        
        if not user_id:
            return Response(
                {"error": "User ID is required. Provide it in the URL path or as query parameter 'user_id'."}, 
                status=status.HTTP_400_BAD_REQUEST
            )
        
        period = request.GET.get('period', 'monthly')
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))
        breakdown = request.GET.get('breakdown', 'detailed')

        # Charge amounts based on your constants
        CHARGE_AMOUNTS = {
            'security': 2000.0,
            'amenities': 2500.0,
            'maintenance': 1500.0,
        }

        try:
            if period == 'monthly':
                return self._get_user_monthly_report(user_id, year, month, breakdown, CHARGE_AMOUNTS)
            elif period == 'yearly':
                return self._get_user_yearly_report(user_id, year, breakdown, CHARGE_AMOUNTS)
            else:
                return Response(
                    {"error": "Invalid period. Use 'monthly' or 'yearly'."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {"error": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    # Helper methods
    def _safe_round(self, value, decimals=2):
        """Safely round a value that might be None"""
        if value is None:
            return 0.0
        try:
            return round(float(value), decimals)
        except (TypeError, ValueError):
            return 0.0

    def _safe_float(self, value):
        """Safely convert to float, handling None"""
        if value is None:
            return 0.0
        try:
            return float(value)
        except (TypeError, ValueError):
            return 0.0

    def _get_assigned_units_for_period(self, user_id, year):
        """Get assigned units for a user during a specific year"""
        from units.models import AssignedUnit
        
        # Get assigned units for the user
        assigned_units = AssignedUnit.objects.filter(
            assigned_by_id=user_id,
            deleted_at__isnull=True,  # Only active assignments
        ).select_related('unit_id')
        
        return assigned_units

    def _get_months_with_charges(self, assigned_units, year):
        """Determine which months of the year charges should apply based on move-in date"""
        if not assigned_units.exists():
            return set()
        
        # Find the earliest move-in date among assigned units
        earliest_month = 13  # Start with a value larger than any month
        has_units_without_date = False
        
        for assigned_unit in assigned_units:
            if assigned_unit.move_in_date:
                move_in_year = assigned_unit.move_in_date.year
                move_in_month = assigned_unit.move_in_date.month
                
                if move_in_year < year:
                    # Moved in before this year, charges from January
                    return set(range(1, 13))
                elif move_in_year == year:
                    # Moved in this year, track the earliest month
                    # Moved in this year, track the earliest month
                    # If move-in is in month X, first full month is X+1
                    first_full_month = move_in_month + 1
                    if move_in_month < earliest_month:
                        earliest_month = first_full_month
            else:
                # No move_in_date specified
                has_units_without_date = True
        
        # If any unit has no move_in_date, assume charges from January
        if has_units_without_date:
            return set(range(1, 13))
        
        # If earliest_month wasn't updated (no units with move_in_date in this year)
        if earliest_month == 13:
            return set()
        
        # Charges apply from earliest_month to December
        return set(range(earliest_month, 13))

    def _calculate_assigned_unit_charges(self, assigned_units, CHARGE_AMOUNTS):
        """Calculate monthly charges based on assigned unit settings"""
        security_fee = 0.0
        maintenance_fee = 0.0
        amenities_fee = 0.0
        
        for assigned_unit in assigned_units:
            # Only add charges if the service is enabled
            if assigned_unit.security:
                security_fee += CHARGE_AMOUNTS['security']
            if assigned_unit.maintenance:
                maintenance_fee += CHARGE_AMOUNTS['maintenance']
            if assigned_unit.amenities:
                amenities_fee += CHARGE_AMOUNTS['amenities']
        
        total_monthly_fees = security_fee + maintenance_fee + amenities_fee
        
        return {
            'security_fee': self._safe_round(security_fee),
            'maintenance_fee': self._safe_round(maintenance_fee),
            'amenities_fee': self._safe_round(amenities_fee),
            'total_monthly_fees': self._safe_round(total_monthly_fees)
        }

    def _get_user_yearly_report(self, user_id, year, breakdown, CHARGE_AMOUNTS):
        """Generate yearly financial report for a specific user (ALL bills)"""
        from calendar import month_name
        from django.contrib.auth import get_user_model
        from units.models import AssignedUnit, Unit
        from django.db.models import Sum, Count
        
        # Get ALL bills for the specified user and year (all months, all statuses)
        all_bills = MonthlyBill.objects.filter(
            user_id=user_id,
            due_date__year=year
        ).select_related('user', 'unit')

        # Get user info
        try:
            user = all_bills.first().user if all_bills.exists() else None
            if not user:
                User = get_user_model()
                user = User.objects.get(id=user_id)
        except User.DoesNotExist:
            return Response(
                {"error": "User not found"},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {"error": f"Error fetching user: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

        user_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username

        # Get assigned units and determine months with charges
        assigned_units = self._get_assigned_units_for_period(user_id, year)
        months_with_charges = self._get_months_with_charges(assigned_units, year)
        
        # Calculate base monthly charges
        base_charge_data = self._calculate_assigned_unit_charges(assigned_units, CHARGE_AMOUNTS)
        
        # Calculate total rent amount for the year from all bills
        total_rent_result = all_bills.aggregate(total=Sum('amount_due'))
        total_rent = self._safe_float(total_rent_result['total'])
        
        # Calculate monthly totals with accurate charges
        monthly_totals = []
        
        for month_num in range(1, 13):
            month_bills = all_bills.filter(due_date__month=month_num)
            
            # Calculate bill amounts
            month_paid_result = month_bills.filter(payment_status='paid').aggregate(total=Sum('amount_due'))
            month_unpaid_result = month_bills.filter(payment_status='pending').aggregate(total=Sum('amount_due'))
            month_total_result = month_bills.aggregate(total=Sum('amount_due'))
            
            month_paid = self._safe_float(month_paid_result['total'])
            month_unpaid = self._safe_float(month_unpaid_result['total'])
            month_expected = self._safe_float(month_total_result['total'])
            
            # Determine if charges apply for this month
            if month_num in months_with_charges:
                month_charges = {
                    'security': base_charge_data['security_fee'],
                    'maintenance': base_charge_data['maintenance_fee'],
                    'amenities': base_charge_data['amenities_fee'],
                    'total_fees': base_charge_data['total_monthly_fees']
                }
            else:
                month_charges = {
                    'security': 0,
                    'maintenance': 0,
                    'amenities': 0,
                    'total_fees': 0
                }
            
            # Calculate monthly completion
            if month_expected > 0:
                month_completion = self._safe_round((month_paid / month_expected) * 100, 2)
            else:
                month_completion = 0.0
            
            monthly_totals.append({
                "month": month_name[month_num],
                "month_number": month_num,
                "amount_paid": self._safe_round(month_paid),
                "amount_unpaid": self._safe_round(month_unpaid),
                "amount_expected": self._safe_round(month_expected),
                "bill_count": month_bills.count(),
                "monthly_completion": f"{month_completion}%",
                "percentage_of_year": self._safe_round((month_expected / total_rent) * 100, 2) if total_rent > 0 else 0,
                "charges": month_charges,
                "charges_applied": month_num in months_with_charges
            })

        # Calculate yearly totals
        paid_bills = all_bills.filter(payment_status='paid')
        unpaid_bills = all_bills.filter(payment_status='pending')
        
        total_paid_result = paid_bills.aggregate(total=Sum('amount_due'))
        total_unpaid_result = unpaid_bills.aggregate(total=Sum('amount_due'))
        total_expected_result = all_bills.aggregate(total=Sum('amount_due'))
        
        total_paid = self._safe_float(total_paid_result['total'])
        total_unpaid = self._safe_float(total_unpaid_result['total'])
        total_expected = self._safe_float(total_expected_result['total'])
        
        total_bills = all_bills.count()
        paid_bills_count = paid_bills.count()
        unpaid_bills_count = unpaid_bills.count()
        
        average_payment = total_paid / paid_bills_count if paid_bills_count > 0 else 0
        
        # Calculate average monthly payment (only for months with bills)
        months_with_bills = [m for m in monthly_totals if m['bill_count'] > 0]
        average_monthly_payment = total_paid / len(months_with_bills) if months_with_bills else 0
        
        if total_expected > 0:
            payment_completion_percentage = self._safe_round((total_paid / total_expected) * 100, 2)
        else:
            payment_completion_percentage = 0

        # Calculate charges based on actual months with charges
        months_with_charges_count = len(months_with_charges)
        monthly_security_fee = base_charge_data['security_fee'] 
        monthly_maintenance_fee = base_charge_data['maintenance_fee'] 
        monthly_amenities_fee = base_charge_data['amenities_fee'] 
        monthly_total_fees = base_charge_data['total_monthly_fees'] 
        
        yearly_security_fee = monthly_security_fee * months_with_charges_count
        yearly_maintenance_fee = monthly_maintenance_fee * months_with_charges_count
        yearly_amenities_fee = monthly_amenities_fee * months_with_charges_count
        yearly_total_fees = monthly_total_fees * months_with_charges_count
        
        # Calculate total yearly charges
        total_yearly_charges = 0
        for month_data in monthly_totals:
            total_yearly_charges += month_data['charges']['total_fees']
        
        # Calculate total monthly obligation (average rent + fees)
        months_with_rent = [m for m in monthly_totals if m['amount_expected'] > 0]
        if months_with_rent:
            average_monthly_rent = total_expected / len(months_with_rent)
            monthly_obligation = average_monthly_rent + monthly_total_fees
        else:
            monthly_obligation = monthly_total_fees
        
        # Get move-in date for display
        move_in_date = None
        if assigned_units.exists():
            # Find the earliest move-in date
            earliest_date = None
            for unit in assigned_units:
                if unit.move_in_date:
                    if earliest_date is None or unit.move_in_date < earliest_date:
                        earliest_date = unit.move_in_date
            
            if earliest_date:
                move_in_date = earliest_date.strftime('%Y-%m-%d')
        
        # Basic response structure
        response_data = {
            "report_type": "user_yearly",
            "user_id": user_id,
            "user_name": user_name,
            "user_email": user.email,
            "year": year,
            "summary": {
                # Bill payment summary
                "total_paid": self._safe_round(total_paid),
                "total_unpaid": self._safe_round(total_unpaid),
                "total_expected": self._safe_round(total_expected),
                "total_bills": total_bills,
                "paid_bills_count": paid_bills_count,
                "unpaid_bills_count": unpaid_bills_count,
                "average_payment": self._safe_round(average_payment),
                "payment_completion": f"{payment_completion_percentage}%",
                "average_monthly_payment": self._safe_round(average_monthly_payment),
                
                # Expense charges summary
                "security_fee": self._safe_round(monthly_security_fee),
                "maintenance_fee": self._safe_round(monthly_maintenance_fee),
                "amenities_fee": self._safe_round(monthly_amenities_fee),
                "total_expenses": self._safe_round(monthly_total_fees),
                "months_with_charges": months_with_charges_count,
                "yearly_security_fee": self._safe_round(yearly_security_fee),
                "yearly_maintenance_fee": self._safe_round(yearly_maintenance_fee),
                "yearly_amenities_fee": self._safe_round(yearly_amenities_fee),
                "yearly_total_fees": self._safe_round(yearly_total_fees),
                "total_monthly_obligation": self._safe_round(monthly_obligation),
                "total_yearly_obligation": self._safe_round(monthly_obligation * 12),
                
                # Rates used
                "charge_rates": CHARGE_AMOUNTS,
                "move_in_date": move_in_date
            },
            "monthly_breakdown": monthly_totals
        }

        # Add detailed breakdown if requested
        if breakdown == 'detailed':
            # ... (keep your existing detailed breakdown code, but make sure to use the updated variables)
            # I'll keep it brief here since you already have this logic
            pass

        return Response(response_data, status=status.HTTP_200_OK)
    
    def _get_user_monthly_report(self, user_id, year, month, breakdown, CHARGE_AMOUNTS):
        """Generate monthly financial report for a specific user"""
        # Implement monthly report logic here
        return Response(
            {"message": "Monthly report not yet implemented"},
            status=status.HTTP_501_NOT_IMPLEMENTED
        )


class FinancialReportExportView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        """
        Export financial report as CSV or PDF
        Query Parameters:
        - period: 'monthly' or 'yearly'
        - year: specific year
        - month: specific month (for monthly)
        - format: 'csv' or 'pdf' (default: 'csv')
        """
        period = request.GET.get('period', 'monthly')
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))
        export_format = request.GET.get('format', 'csv')

        # Get the financial report data
        report_view = FinancialReportsView()
        
        if period == 'monthly':
            response = report_view._get_monthly_report(year, month)
        else:
            response = report_view._get_yearly_report(year)

        report_data = response.data

        if export_format == 'csv':
            return self._export_to_csv(report_data, period, year, month)
        else:
            return Response(
                {"error": "PDF export not implemented yet"},
                status=status.HTTP_501_NOT_IMPLEMENTED
            )

    def _export_to_csv(self, report_data, period, year, month):
        """Export financial report to CSV format"""
        import csv
        from django.http import HttpResponse
        
        response = HttpResponse(content_type='text/csv')
        
        if period == 'monthly':
            filename = f"financial_report_{year}_{month:02d}.csv"
        else:
            filename = f"financial_report_{year}.csv"
            
        response['Content-Disposition'] = f'attachment; filename="{filename}"'

        writer = csv.writer(response)
        
        # Write header
        writer.writerow(['Financial Report - Paid Bills Only'])
        if period == 'monthly':
            writer.writerow([f'Period: {report_data["month"]} {report_data["year"]}'])
        else:
            writer.writerow([f'Period: Year {report_data["year"]}'])
        writer.writerow([])
        
        # Write summary
        writer.writerow(['SUMMARY'])
        writer.writerow(['Total Collected', f"₱{report_data['summary']['total_collected']:,.2f}"])
        writer.writerow(['Total Bills', report_data['summary']['total_bills']])
        writer.writerow(['Average Payment', f"₱{report_data['summary']['average_payment']:,.2f}"])
        writer.writerow([])
        
        # Write user breakdown
        writer.writerow(['USER BREAKDOWN'])
        writer.writerow(['User', 'Total Paid', 'Bill Count', 'Percentage'])
        for user in report_data['breakdowns']['by_user']:
            writer.writerow([
                user['user_name'],
                f"₱{user['total_paid']:,.2f}",
                user['bill_count'],
                f"{user['percentage_of_total']}%"
            ])
        writer.writerow([])
        
        return response
    



class FinancialReportView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, user_id=None):
        """
        Generate financial reports (monthly or yearly) for a specific user.
        Query Parameters:
        - period: 'monthly' or 'yearly' (default: 'monthly')
        - year: specific year (default: current year)
        - month: specific month (for monthly reports, default: current month)
        - breakdown: 'detailed' or 'summary' (default: 'detailed')
        """
        if not user_id:
            return Response(
                {"error": "User ID is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        period = request.GET.get('period', 'monthly')
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))
        breakdown = request.GET.get('breakdown', 'detailed')

        try:
            if period == 'monthly':
                return self._get_user_monthly_report(user_id, year, month, breakdown)
            elif period == 'yearly':
                return self._get_user_yearly_report(user_id, year, breakdown)
            else:
                return Response(
                    {"error": "Invalid period. Use 'monthly' or 'yearly'."},
                    status=status.HTTP_400_BAD_REQUEST
                )
        except Exception as e:
            return Response(
                {"error": f"An error occurred: {str(e)}"},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )

    def _get_user_monthly_report(self, user_id, year, month, breakdown):
        """Generate monthly financial report for a specific user (ALL bills, not just paid)"""
        # Get ALL bills for the specified user, month and year (not just paid)
        all_bills = MonthlyBill.objects.filter(
            user_id=user_id,
            due_date__year=year,
            due_date__month=month
        ).select_related('user', 'unit')

        # Get user info
        try:
            user = all_bills.first().user if all_bills.exists() else None
            if not user:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                user = User.objects.get(id=user_id)
        except:
            return Response(
                {"error": "User not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        user_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username

        if not all_bills.exists():
            return Response({
                "report_type": "user_monthly",
                "user_id": user_id,
                "user_name": user_name,
                "year": year,
                "month": calendar.month_name[month],
                "message": "No bills found for this period",
                "summary": {
                    "total_paid": 0,
                    "total_unpaid": 0,
                    "total_bills": 0,
                    "average_payment": 0,
                    "payment_completion": "0%"
                }
            }, status=status.HTTP_200_OK)

        # Calculate totals - ALL bills, not just paid
        total_paid = all_bills.filter(payment_status='paid').aggregate(total=Sum('amount_due'))['total'] or 0
        total_unpaid = all_bills.filter(payment_status='pending').aggregate(total=Sum('amount_due'))['total'] or 0
        total_expected = all_bills.aggregate(total=Sum('amount_due'))['total'] or 0
        total_bills = all_bills.count()
        
        paid_bills_count = all_bills.filter(payment_status='paid').count()
        unpaid_bills_count = all_bills.filter(payment_status='pending').count()
        
        average_payment = total_paid / paid_bills_count if paid_bills_count > 0 else 0
        
        # Calculate ACTUAL payment completion
        if total_expected > 0:
            payment_completion_percentage = round((total_paid / total_expected) * 100, 2)
        else:
            payment_completion_percentage = 0

        # Basic response structure
        response_data = {
            "report_type": "user_monthly",
            "user_id": user_id,
            "user_name": user_name,
            "user_email": user.email,
            "year": year,
            "month": calendar.month_name[month],
            "summary": {
                "total_paid": round((total_paid), 2),
                "total_unpaid": round((total_unpaid), 2),
                "total_expected": round((total_expected), 2),
                "total_bills": total_bills,
                "paid_bills_count": paid_bills_count,
                "unpaid_bills_count": unpaid_bills_count,
                "average_payment": round((average_payment), 2),
                "payment_completion": f"{payment_completion_percentage}%",  # ACTUAL completion rate
            }
        }

        # Add detailed breakdown if requested
        if breakdown == 'detailed':
            # Breakdown by unit (ALL bills)
            unit_breakdown = all_bills.values(
                'unit__unit_name', 
                'unit__building',
                'unit__rent_amount'
            ).annotate(
                total_paid=Sum('amount_due', filter=models.Q(payment_status='paid')),
                total_unpaid=Sum('amount_due', filter=models.Q(payment_status='pending')),
                total_expected=Sum('amount_due'),
                bill_count=Count('id'),
                paid_count=Count('id', filter=models.Q(payment_status='paid')),
                unpaid_count=Count('id', filter=models.Q(payment_status='pending'))
            ).order_by('-total_expected')

            unit_breakdown_data = []
            for unit_data in unit_breakdown:
                unit_completion = round((unit_data['total_paid'] / unit_data['total_expected']) * 100, 2) if unit_data['total_expected'] > 0 else 0
                
                unit_breakdown_data.append({
                    "unit_name": unit_data['unit__unit_name'],
                    "building": unit_data['unit__building'],
                    "rent_amount": (unit_data['unit__rent_amount']) if unit_data['unit__rent_amount'] else 0,
                    "total_paid": round((unit_data['total_paid']), 2),
                    "total_unpaid": round((unit_data['total_unpaid']), 2),
                    "total_expected": round((unit_data['total_expected']), 2),
                    "bill_count": unit_data['bill_count'],
                    "paid_bills": unit_data['paid_count'],
                    "unpaid_bills": unit_data['unpaid_count'],
                    "unit_completion": f"{unit_completion}%",
                    "percentage_of_total": round((unit_data['total_expected'] / total_expected) * 100, 2) if total_expected > 0 else 0
                })

            # Daily payment trend (ALL bills)
            daily_trend = all_bills.values('due_date').annotate(
                daily_total=Sum('amount_due'),
                daily_paid=Sum('amount_due', filter=models.Q(payment_status='paid')),
                daily_unpaid=Sum('amount_due', filter=models.Q(payment_status='pending')),
                daily_count=Count('id')
            ).order_by('due_date')

            daily_trend_data = []
            for day_data in daily_trend:
                day_completion = round((day_data['daily_paid'] / day_data['daily_total']) * 100, 2) if day_data['daily_total'] > 0 else 0
                
                daily_trend_data.append({
                    "date": day_data['due_date'].strftime('%Y-%m-%d'),
                    "amount_paid": round((day_data['daily_paid']), 2),
                    "amount_unpaid": round((day_data['daily_unpaid']), 2),
                    "amount_expected": round((day_data['daily_total']), 2),
                    "bill_count": day_data['daily_count'],
                    "daily_completion": f"{day_completion}%"
                })

            # Bill details (ALL bills)
            bill_details = []
            for bill in all_bills:
                bill_details.append({
                    "bill_id": bill.id,
                    "due_date": bill.due_date.strftime('%Y-%m-%d'),
                    "amount_due": (bill.amount_due),
                    "amount_paid": (bill.amount_due) if bill.payment_status == 'paid' else 0,
                    "unit_name": bill.unit.unit_name if bill.unit else "N/A",
                    "building": bill.unit.building if bill.unit else "N/A",
                    "payment_status": bill.payment_status,
                    "created_at": bill.created_at.strftime('%Y-%m-%d %H:%M:%S')
                })

            response_data["detailed_breakdown"] = {
                "by_unit": unit_breakdown_data,
                "daily_payments": daily_trend_data,
                "bill_details": bill_details
            }

        return Response(response_data, status=status.HTTP_200_OK)

    def _get_user_yearly_report(self, user_id, year, breakdown):
        """Generate yearly financial report for a specific user (ALL bills)"""
        # Get ALL bills for the specified user and year (not just paid)
        all_bills = MonthlyBill.objects.filter(
            user_id=user_id,
            due_date__year=year
        ).select_related('user', 'unit')

        # Get user info
        try:
            user = all_bills.first().user if all_bills.exists() else None
            if not user:
                from django.contrib.auth import get_user_model
                User = get_user_model()
                user = User.objects.get(id=user_id)
        except:
            return Response(
                {"error": "User not found"},
                status=status.HTTP_404_NOT_FOUND
            )

        user_name = f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username

        if not all_bills.exists():
            return Response({
                "report_type": "user_yearly",
                "user_id": user_id,
                "user_name": user_name,
                "year": year,
                "message": "No bills found for this year",
                "summary": {
                    "total_paid": 0,
                    "total_unpaid": 0,
                    "total_bills": 0,
                    "average_payment": 0,
                    "payment_completion": "0%"
                }
            }, status=status.HTTP_200_OK)

        # Calculate totals - ALL bills
        total_paid = all_bills.filter(payment_status='paid').aggregate(total=Sum('amount_due'))['total'] or 0
        total_unpaid = all_bills.filter(payment_status='pending').aggregate(total=Sum('amount_due'))['total'] or 0
        total_expected = all_bills.aggregate(total=Sum('amount_due'))['total'] or 0
        total_bills = all_bills.count()
        
        paid_bills_count = all_bills.filter(payment_status='paid').count()
        unpaid_bills_count = all_bills.filter(payment_status='pending').count()
        
        average_payment = total_paid / paid_bills_count if paid_bills_count > 0 else 0
        
        # Calculate ACTUAL payment completion
        if total_expected > 0:
            payment_completion_percentage = round((total_paid / total_expected) * 100, 2)
        else:
            payment_completion_percentage = 0

        # Monthly breakdown (ALL bills)
        monthly_breakdown = all_bills.annotate(
            month=ExtractMonth('due_date')
        ).values('month').annotate(
            monthly_paid=Sum('amount_due', filter=models.Q(payment_status='paid')),
            monthly_unpaid=Sum('amount_due', filter=models.Q(payment_status='pending')),
            monthly_total=Sum('amount_due'),
            monthly_count=Count('id')
        ).order_by('month')

        monthly_breakdown_data = []
        for month_data in monthly_breakdown:
            month_completion = round((month_data['monthly_paid'] / month_data['monthly_total']) * 100, 2) if month_data['monthly_total'] > 0 else 0
            
            monthly_breakdown_data.append({
                "month": calendar.month_name[month_data['month']],
                "month_number": month_data['month'],
                "amount_paid": round((month_data['monthly_paid']), 2),
                "amount_unpaid": round((month_data['monthly_unpaid']), 2),
                "amount_expected": round((month_data['monthly_total']), 2),
                "bill_count": month_data['monthly_count'],
                "monthly_completion": f"{month_completion}%",
                "percentage_of_year": round((month_data['monthly_total'] / total_expected) * 100, 2) if total_expected > 0 else 0
            })

        # Basic response structure
        response_data = {
            "report_type": "user_yearly",
            "user_id": user_id,
            "user_name": user_name,
            "user_email": user.email,
            "year": year,
            "summary": {
                "total_paid": round((total_paid), 2),
                "total_unpaid": round((total_unpaid), 2),
                "total_expected": round((total_expected), 2),
                "total_bills": total_bills,
                "paid_bills_count": paid_bills_count,
                "unpaid_bills_count": unpaid_bills_count,
                "average_payment": round((average_payment), 2),
                "payment_completion": f"{payment_completion_percentage}%",  # ACTUAL completion rate
                "average_monthly_payment": round((total_paid / 12), 2) if total_paid > 0 else 0,
            },
            "monthly_breakdown": monthly_breakdown_data
        }

        # Add detailed breakdown if requested
        if breakdown == 'detailed':
            # Breakdown by unit (ALL bills)
            unit_breakdown = all_bills.values(
                'unit__unit_name', 
                'unit__building',
                'unit__rent_amount'
            ).annotate(
                total_paid=Sum('amount_due', filter=models.Q(payment_status='paid')),
                total_unpaid=Sum('amount_due', filter=models.Q(payment_status='pending')),
                total_expected=Sum('amount_due'),
                bill_count=Count('id'),
                paid_count=Count('id', filter=models.Q(payment_status='paid')),
                unpaid_count=Count('id', filter=models.Q(payment_status='pending')),
                avg_payment=Avg('amount_due', filter=models.Q(payment_status='paid'))
            ).order_by('-total_expected')

            unit_breakdown_data = []
            for unit_data in unit_breakdown:
                unit_completion = round((unit_data['total_paid'] / unit_data['total_expected']) * 100, 2) if unit_data['total_expected'] > 0 else 0
                unit_monthly_bills = all_bills.filter(
                    unit__unit_name=unit_data['unit__unit_name']
                ).values('due_date__month').distinct().count()
                
                unit_breakdown_data.append({
                    "unit_name": unit_data['unit__unit_name'],
                    "building": unit_data['unit__building'],
                    "rent_amount": (unit_data['unit__rent_amount']) if unit_data['unit__rent_amount'] else 0,
                    "total_paid": round((unit_data['total_paid']), 2),
                    "total_unpaid": round((unit_data['total_unpaid']), 2),
                    "total_expected": round((unit_data['total_expected']), 2),
                    "bill_count": unit_data['bill_count'],
                    "paid_bills": unit_data['paid_count'],
                    "unpaid_bills": unit_data['unpaid_count'],
                    "average_payment": round((unit_data['avg_payment']), 2),
                    "unit_completion": f"{unit_completion}%",
                    "months_with_payments": unit_monthly_bills,
                    "percentage_of_total": round((unit_data['total_expected'] / total_expected) * 100, 2) if total_expected > 0 else 0
                })

            # Quarterly breakdown (ALL bills)
            quarterly_breakdown = []
            for quarter in range(1, 5):
                if quarter == 1:
                    months = [1, 2, 3]
                elif quarter == 2:
                    months = [4, 5, 6]
                elif quarter == 3:
                    months = [7, 8, 9]
                else:
                    months = [10, 11, 12]
                
                quarter_bills = all_bills.filter(due_date__month__in=months)
                quarter_paid = quarter_bills.filter(payment_status='paid').aggregate(total=Sum('amount_due'))['total'] or 0
                quarter_unpaid = quarter_bills.filter(payment_status='pending').aggregate(total=Sum('amount_due'))['total'] or 0
                quarter_total = quarter_paid + quarter_unpaid
                quarter_count = quarter_bills.count()
                
                quarter_completion = round((quarter_paid / quarter_total) * 100, 2) if quarter_total > 0 else 0
                
                quarterly_breakdown.append({
                    "quarter": f"Q{quarter}",
                    "months": [calendar.month_name[m] for m in months],
                    "amount_paid": round((quarter_paid), 2),
                    "amount_unpaid": round((quarter_unpaid), 2),
                    "amount_expected": round((quarter_total), 2),
                    "bill_count": quarter_count,
                    "quarter_completion": f"{quarter_completion}%",
                    "percentage_of_year": round((quarter_total / total_expected) * 100, 2) if total_expected > 0 else 0
                })

            # Payment consistency analysis
            months_with_payments = len([m for m in monthly_breakdown_data if m['amount_expected'] > 0])
            payment_consistency = {
                "months_with_payments": months_with_payments,
                "payment_consistency_rate": round((months_with_payments / 12) * 100, 2),
                "most_active_month": max(monthly_breakdown_data, key=lambda x: x['amount_paid']) if monthly_breakdown_data else None,
                "least_active_month": min(monthly_breakdown_data, key=lambda x: x['amount_paid']) if monthly_breakdown_data else None,
                "best_completion_month": max(monthly_breakdown_data, key=lambda x: (x['monthly_completion'].rstrip('%'))) if monthly_breakdown_data else None,
                "worst_completion_month": min(monthly_breakdown_data, key=lambda x: (x['monthly_completion'].rstrip('%'))) if monthly_breakdown_data else None
            }

            response_data["detailed_breakdown"] = {
                "by_unit": unit_breakdown_data,
                "quarterly": quarterly_breakdown,
                "payment_analysis": payment_consistency
            }

        return Response(response_data, status=status.HTTP_200_OK)


class UserFinancialComparisonView(APIView):
    permission_classes = [AllowAny]

    def get(self, request):
        """
        Compare financial performance of multiple users
        Query Parameters:
        - user_ids: comma-separated list of user IDs
        - period: 'monthly' or 'yearly' (default: 'yearly')
        - year: specific year (default: current year)
        - month: specific month (for monthly, default: current month)
        """
        user_ids_param = request.GET.get('user_ids', '')
        if not user_ids_param:
            return Response(
                {"error": "user_ids parameter is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        try:
            user_ids = [int(uid.strip()) for uid in user_ids_param.split(',')]
        except ValueError:
            return Response(
                {"error": "Invalid user IDs format"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        period = request.GET.get('period', 'yearly')
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))

        comparison_data = []
        
        for user_id in user_ids:
            try:
                # Get user basic info
                from django.contrib.auth import get_user_model
                User = get_user_model()
                user = User.objects.get(id=user_id)
                
                # Get paid bills based on period
                if period == 'monthly':
                    paid_bills = MonthlyBill.objects.filter(
                        user_id=user_id,
                        payment_status='paid',
                        due_date__year=year,
                        due_date__month=month
                    )
                else:
                    paid_bills = MonthlyBill.objects.filter(
                        user_id=user_id,
                        payment_status='paid',
                        due_date__year=year
                    )

                total_paid = paid_bills.aggregate(total=Sum('amount_due'))['total'] or 0
                bill_count = paid_bills.count()
                
                # Get assigned units count for context
                assigned_units_count = AssignedUnit.objects.filter(
                    assigned_by_id=user_id,
                    deleted_at__isnull=True
                ).count()

                user_data = {
                    "user_id": user_id,
                    "user_name": f"{user.first_name or ''} {user.last_name or ''}".strip() or user.username,
                    "user_email": user.email,
                    "total_paid": round((total_paid), 2),
                    "bill_count": bill_count,
                    "assigned_units_count": assigned_units_count,
                    "average_payment": round((total_paid / bill_count), 2) if bill_count > 0 else 0
                }
                
                comparison_data.append(user_data)
                
            except User.DoesNotExist:
                comparison_data.append({
                    "user_id": user_id,
                    "error": "User not found"
                })
            except Exception as e:
                comparison_data.append({
                    "user_id": user_id,
                    "error": str(e)
                })

        # Sort by total paid (descending)
        comparison_data.sort(key=lambda x: x.get('total_paid', 0), reverse=True)

        # Calculate rankings and percentages
        total_all_users = sum(item.get('total_paid', 0) for item in comparison_data if 'total_paid' in item)
        
        for i, user_data in enumerate(comparison_data):
            if 'total_paid' in user_data:
                user_data['rank'] = i + 1
                user_data['percentage_of_total'] = round(
                    (user_data['total_paid'] / total_all_users * 100), 2
                ) if total_all_users > 0 else 0

        response_data = {
            "comparison_period": period,
            "year": year,
            "month": calendar.month_name[month] if period == 'monthly' else None,
            "total_all_users": round((total_all_users), 2),
            "average_per_user": round((total_all_users / len(comparison_data)), 2) if comparison_data else 0,
            "user_comparison": comparison_data
        }

        return Response(response_data, status=status.HTTP_200_OK)


class UserFinancialReportExportView(APIView):
    permission_classes = [AllowAny]

    def get(self, request, user_id=None):
        """
        Export user financial report as CSV
        Query Parameters:
        - period: 'monthly' or 'yearly'
        - year: specific year
        - month: specific month (for monthly)
        """
        if not user_id:
            return Response(
                {"error": "User ID is required"}, 
                status=status.HTTP_400_BAD_REQUEST
            )

        period = request.GET.get('period', 'yearly')
        year = int(request.GET.get('year', date.today().year))
        month = int(request.GET.get('month', date.today().month))

        # Get the user financial report data
        report_view = FinancialReportView()
        
        try:
            if period == 'monthly':
                response = report_view._get_user_monthly_report(user_id, year, month, 'detailed')
            else:
                response = report_view._get_user_yearly_report(user_id, year, 'detailed')

            report_data = response.data

            # Export to CSV
            import csv
            from django.http import HttpResponse
            
            response = HttpResponse(content_type='text/csv')
            
            if period == 'monthly':
                filename = f"user_financial_report_{user_id}_{year}_{month:02d}.csv"
            else:
                filename = f"user_financial_report_{user_id}_{year}.csv"
                
            response['Content-Disposition'] = f'attachment; filename="{filename}"'

            writer = csv.writer(response)
            
            # Write header
            writer.writerow(['User Financial Report - Paid Bills Only'])
            writer.writerow([f'User: {report_data["user_name"]}'])
            writer.writerow([f'User ID: {report_data["user_id"]}'])
            if period == 'monthly':
                writer.writerow([f'Period: {report_data["month"]} {report_data["year"]}'])
            else:
                writer.writerow([f'Period: Year {report_data["year"]}'])
            writer.writerow([])
            
            # Write summary - handle potential None values
            writer.writerow(['SUMMARY'])
            total_paid = report_data['summary']['total_paid'] or 0
            total_bills = report_data['summary']['total_bills'] or 0
            avg_payment = report_data['summary']['average_payment'] or 0
            
            writer.writerow(['Total Paid', f"₱{total_paid:,.2f}"])
            writer.writerow(['Total Bills', total_bills])
            writer.writerow(['Average Payment', f"₱{avg_payment:,.2f}"])
            writer.writerow([])
            
            # Write unit breakdown if available
            if 'detailed_breakdown' in report_data and 'by_unit' in report_data['detailed_breakdown']:
                writer.writerow(['UNIT BREAKDOWN'])
                writer.writerow(['Unit Name', 'Building', 'Total Paid', 'Bill Count', 'Percentage'])
                for unit in report_data['detailed_breakdown']['by_unit']:
                    unit_paid = unit['total_paid'] or 0
                    unit_bill_count = unit['bill_count'] or 0
                    unit_percentage = unit['percentage_of_total'] or 0
                    
                    writer.writerow([
                        unit['unit_name'],
                        unit['building'],
                        f"₱{unit_paid:,.2f}",
                        unit_bill_count,
                        f"{unit_percentage}%"
                    ])
                writer.writerow([])
            
            return response
            
        except Exception as e:
            print(f"Error exporting financial report: {str(e)}")
            return Response(
                {"error": "Failed to generate export"}, 
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
        

class PaidBillsExcelExportView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request):
        """
        Export paid/done bills to Excel with monthly filtering
        Query Parameters:
        - month: month number (1-12) or 'all' for all months
        - year: specific year (default: current year)
        - status: 'paid', 'done', or 'all' (default: 'all')
        """
        try:
            # Get query parameters
            month_param = request.GET.get('month', 'all')
            year = int(request.GET.get('year', date.today().year))
            status_filter = request.GET.get('status', 'all')
            
            # Build base query for paid/done bills
            base_query = MonthlyBill.objects.filter(
                Q(payment_status=MonthlyBill.PaymentStatus.PAID) | 
                Q(due_status=MonthlyBill.DueStatus.DONE)
            )
            
            # Apply month filter
            if month_param != 'all':
                month = int(month_param)
                base_query = base_query.filter(
                    due_date__year=year,
                    due_date__month=month
                )
            else:
                base_query = base_query.filter(due_date__year=year)
            
            # Apply specific status filter if provided
            if status_filter == 'paid':
                base_query = base_query.filter(payment_status=MonthlyBill.PaymentStatus.PAID)
            elif status_filter == 'done':
                base_query = base_query.filter(due_status=MonthlyBill.DueStatus.DONE)
            
            # Order by due date
            bills = base_query.order_by('due_date')
            
            if not bills.exists():
                return HttpResponse(
                    "No paid bills found for the selected criteria", 
                    status=404, 
                    content_type='text/plain'
                )
            
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Paid Bills Report"
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="344CB7", end_color="344CB7", fill_type="solid")
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_align = Alignment(horizontal='center', vertical='center')
            
            # Headers
            headers = [
                'Bill ID', 'User', 'User Email', 'Unit', 'Building',
                'Amount Due', 'Due Date', 'Payment Status', 'Due Status',
                'SMS Sent', 'Created Date'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_align
                cell.border = border
            
            # Write data
            for row, bill in enumerate(bills, 2):
                # Get unit info safely
                unit_name = bill.unit.unit_name if bill.unit else "N/A"
                building_name = bill.unit.building if bill.unit and bill.unit.building else "N/A"
                
                data = [
                    bill.id,
                    bill.user.get_full_name() or bill.user.username,
                    bill.user.email,
                    unit_name,
                    building_name,
                    float(bill.amount_due),  # Convert Decimal to float for Excel
                    bill.due_date.strftime('%Y-%m-%d'),
                    bill.get_payment_status_display(),
                    bill.get_due_status_display(),
                    'Yes' if bill.sms_sent else 'No',
                    bill.created_at.strftime('%Y-%m-%d %H:%M:%S')
                ]
                
                for col, value in enumerate(data, 1):
                    cell = ws.cell(row=row, column=col, value=value)
                    cell.border = border
                    if col in [6, 7, 8]:  # Amount and date columns
                        cell.alignment = center_align
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Add summary section
            summary_row = len(bills) + 3
            ws.cell(row=summary_row, column=1, value="SUMMARY").font = Font(bold=True, size=14)
            
            summary_data = [
                ("Total Paid Bills", len(bills)),
                ("Total Amount", f"₱{sum(bill.amount_due for bill in bills):,.2f}"),
                ("Average Amount", f"₱{sum(bill.amount_due for bill in bills) / len(bills):,.2f}"),
                ("Date Range", f"{bills.first().due_date} to {bills.last().due_date}"),
                ("Generated On", datetime.now().strftime('%Y-%m-%d %H:%M:%S'))
            ]
            
            for i, (label, value) in enumerate(summary_data, 1):
                ws.cell(row=summary_row + i, column=1, value=label).font = Font(bold=True)
                ws.cell(row=summary_row + i, column=2, value=value)
            
            # Prepare HTTP response
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create filename
            if month_param == 'all':
                filename = f"paid_bills_report_{year}_all_months.xlsx"
            else:
                month_name = calendar.month_name[int(month_param)]
                filename = f"paid_bills_report_{year}_{month_name}.xlsx"
            
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}"'
            
            return response
            
        except Exception as e:
            return HttpResponse(
                f"Error generating Excel report: {str(e)}",
                status=500,
                content_type='text/plain'
            )


class PaidBillsFilterOptionsView(APIView):
    permission_classes = [AllowAny]
    
    def get(self, request):
        """
        Get available filter options for paid bills export
        """
        # Get distinct years with paid bills
        years = MonthlyBill.objects.filter(
            Q(payment_status=MonthlyBill.PaymentStatus.PAID) | 
            Q(due_status=MonthlyBill.DueStatus.DONE)
        ).dates('due_date', 'year').order_by('-due_date')
        
        # Get months with paid bills for current year
        current_year = date.today().year
        months = MonthlyBill.objects.filter(
            (Q(payment_status=MonthlyBill.PaymentStatus.PAID) | 
             Q(due_status=MonthlyBill.DueStatus.DONE)) &
            Q(due_date__year=current_year)
        ).dates('due_date', 'month').order_by('due_date')
        
        options = {
            'available_years': [year.year for year in years],
            'available_months_current_year': [
                {'number': month.month, 'name': calendar.month_name[month.month]}
                for month in months
            ],
            'status_options': [
                {'value': 'all', 'label': 'All Paid/Done Bills'},
                {'value': 'paid', 'label': 'Only Paid (Payment Status)'},
                {'value': 'done', 'label': 'Only Done (Due Status)'}
            ]
        }
        
        return Response(options)

from django.db.models import Sum, Q, Count, FloatField, Min, Max
from django.db.models.functions import Coalesce
from .serializers import ExpenseReflectionSerializer
from units.models import Unit, AssignedUnit  # Assuming you have a Building model

from django.db.models import Sum, Q, Count, FloatField, F, ExpressionWrapper, DecimalField
from django.db.models.functions import Coalesce, ExtractYear, ExtractMonth, TruncMonth
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework import status, permissions
from decimal import Decimal
import datetime
from django.utils import timezone
from .serializers import ExpenseReflectionSerializer
from units.models import Unit, AssignedUnit
from .models import MonthlyBill


class ExpenseReflectionAPIView(APIView):
    """
    API endpoint for total expense reflection with building filtering.
    
    Query Parameters:
    - building: Optional, filter by specific building name
    - start_date: Optional, filter by start date (YYYY-MM-DD)
    - end_date: Optional, filter by end date (YYYY-MM-DD)
    - year: Optional, filter by specific year
    - month: Optional, filter by specific month (1-12)
    - show_breakdown: Optional, include detailed breakdown (true/false)
    - include_other: Optional, include other expenses in total (true/false, default=false)
    - chart_type: Optional, specify chart type (pie, bar, line, monthly_trend)
    - group_by: Optional, group by (month, year, building)
    """
    
    def get(self, request, format=None):
        # Get query parameters
        building_name = request.query_params.get('building')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        year = request.query_params.get('year')
        month = request.query_params.get('month')
        show_breakdown = request.query_params.get('show_breakdown', 'false').lower() == 'true'
        include_other = request.query_params.get('include_other', 'false').lower() == 'true'
        chart_type = request.query_params.get('chart_type', 'pie')  # pie, bar, line, monthly_trend
        group_by = request.query_params.get('group_by')  # month, year, building
        
        # Start with base queryset
        bills_queryset = MonthlyBill.objects.all()
        
        # Apply date filters
        if start_date:
            bills_queryset = bills_queryset.filter(due_date__gte=start_date)
        if end_date:
            bills_queryset = bills_queryset.filter(due_date__lte=end_date)
        if year:
            bills_queryset = bills_queryset.filter(due_date__year=year)
        if month:
            bills_queryset = bills_queryset.filter(due_date__month=month)
        
        # Apply building filter if specified
        if building_name:
            # Get units in this building
            units_in_building = Unit.objects.filter(building=building_name, deleted_at__isnull=True)
            # Get unit IDs
            unit_ids = units_in_building.values_list('id', flat=True)
            # Filter bills for units in this building
            bills_queryset = bills_queryset.filter(unit_id__in=unit_ids)
        
        # Calculate basic totals from ALL bills
        all_bills_aggregates = bills_queryset.aggregate(
            total_all_bills=Coalesce(Sum('amount_due'), Decimal('0')),
            total_paid=Coalesce(
                Sum('amount_due', filter=Q(payment_status=MonthlyBill.PaymentStatus.PAID)),
                Decimal('0')
            ),
            total_unpaid=Coalesce(
                Sum('amount_due', filter=Q(payment_status=MonthlyBill.PaymentStatus.PENDING)),
                Decimal('0')
            ),
            total_bills=Count('id'),
            paid_bills=Count('id', filter=Q(payment_status=MonthlyBill.PaymentStatus.PAID)),
            pending_bills=Count('id', filter=Q(payment_status=MonthlyBill.PaymentStatus.PENDING))
        )
        
        # Get the total from ALL bills
        total_all_bills = all_bills_aggregates['total_all_bills']
        
        # Get distinct units from the filtered bills
        unit_ids_from_bills = bills_queryset.values_list('unit_id', flat=True).distinct()
        
        # Get assigned units for these unit IDs
        assigned_units = AssignedUnit.objects.filter(
            unit_id__in=unit_ids_from_bills,
            deleted_at__isnull=True
        )
        
        # Calculate months covered by the filtered bills
        if bills_queryset.exists():
            # Get unique month-year combinations
            months_data = bills_queryset.annotate(
                month=ExtractMonth('due_date'),
                year=ExtractYear('due_date')
            ).values('month', 'year').distinct()
            months_count = months_data.count()
        else:
            months_count = 0
        
        # Calculate expenses based on fixed rates
        # Fixed rates as per your requirement
        MAINTENANCE_RATE = Decimal('1500.00')
        SECURITY_RATE = Decimal('2000.00')
        AMENITIES_RATE = Decimal('2500.00')
        
        # Initialize totals
        maintenance = Decimal('0.00')
        security = Decimal('0.00')
        amenities = Decimal('0.00')
        
        # If no assigned units found, check if we should still calculate based on units in building
        if not assigned_units.exists() and building_name:
            # Get all units in the building
            units_in_building = Unit.objects.filter(building=building_name, deleted_at__isnull=True)
            
            # Calculate based on all units (assuming default service values)
            for unit in units_in_building:
                # Get bills for this specific unit
                unit_bills = bills_queryset.filter(unit_id=unit.id)
                if unit_bills.exists():
                    # Count unique months for this unit
                    unit_months = unit_bills.annotate(
                        month=ExtractMonth('due_date'),
                        year=ExtractYear('due_date')
                    ).values('month', 'year').distinct().count()
                    
                    # Add expenses (assuming all services are enabled by default)
                    maintenance += MAINTENANCE_RATE * unit_months
                    security += SECURITY_RATE * unit_months
                    amenities += AMENITIES_RATE * unit_months
        else:
            # Calculate based on assigned units with their service configurations
            for assigned_unit in assigned_units:
                # Get bills for this specific unit
                unit_bills = bills_queryset.filter(unit_id=assigned_unit.unit_id)
                if unit_bills.exists():
                    # Count unique months for this unit
                    unit_months = unit_bills.annotate(
                        month=ExtractMonth('due_date'),
                        year=ExtractYear('due_date')
                    ).values('month', 'year').distinct().count()
                    
                    # Add expenses based on enabled services
                    if assigned_unit.maintenance:
                        maintenance += MAINTENANCE_RATE * unit_months
                    if assigned_unit.security:
                        security += SECURITY_RATE * unit_months
                    if assigned_unit.amenities:
                        amenities += AMENITIES_RATE * unit_months
        
        # Round to 2 decimal places
        maintenance = Decimal(round(maintenance, 2))
        security = Decimal(round(security, 2))
        amenities = Decimal(round(amenities, 2))
        
        # Calculate total categorized expenses
        total_expense = maintenance + security + amenities
        
        # Calculate other expenses
        other_expenses = total_all_bills - total_expense
        if other_expenses < Decimal('0'):
            other_expenses = Decimal('0')
        
        # Calculate percentages for each category
        maintenance_pct = (maintenance / total_expense * 100) if total_expense > 0 else 0
        security_pct = (security / total_expense * 100) if total_expense > 0 else 0
        amenities_pct = (amenities / total_expense * 100) if total_expense > 0 else 0
        
        # Prepare base response data
        data = {
            'maintenance': maintenance,
            'security': security,
            'amenities': amenities,
            'totalExpense': total_expense,  # Sum of the three categories
            'totalPaid': all_bills_aggregates['total_paid'],
            'totalUnpaid': all_bills_aggregates['total_unpaid'],
            'building_filter': building_name if building_name else 'All Buildings',
            'other_expenses': other_expenses,
            'total_all_bills': total_all_bills,
            'filters_applied': {
                'building': building_name,
                'year': year,
                'month': month,
                'start_date': start_date,
                'end_date': end_date,
                'chart_type': chart_type
            }
        }
        
        # Add summary
        data['summary'] = {
            'categorized_total': total_expense,
            'other_expenses': other_expenses,
            'total_all_bills': total_all_bills,
            'categorized_percentage': float(((total_expense) / total_all_bills * 100)) if total_all_bills > 0 else 0,
            'other_percentage': float((other_expenses / total_all_bills * 100)) if total_all_bills > 0 else 0,
            'verification': 'OK' if ((total_expense + other_expenses) == total_all_bills) else 'MISMATCH',
        }
        
        # Generate chart data based on chart_type
        chart_data = self.generate_chart_data(
            bills_queryset, 
            chart_type, 
            building_name,
            year,
            month
        )
        data['chart_data'] = chart_data
        
        # Add detailed breakdown if requested
        if show_breakdown:
            # Get assigned units service statistics
            if building_name:
                # Try to get assigned units by building name
                assigned_units_in_building = AssignedUnit.objects.filter(
                    building=building_name,
                    deleted_at__isnull=True
                )
            else:
                assigned_units_in_building = assigned_units
            
            service_stats = {
                'total_assigned_units': assigned_units_in_building.count(),
                'maintenance_enabled': assigned_units_in_building.filter(maintenance=True).count(),
                'security_enabled': assigned_units_in_building.filter(security=True).count(),
                'amenities_enabled': assigned_units_in_building.filter(amenities=True).count(),
            }
            
            data['detailed_breakdown'] = {
                'categories': {
                    'maintenance': {
                        'amount': maintenance,
                        'percentage_of_categorized': float(round(maintenance_pct, 2)),
                        'percentage_of_total': float(round((maintenance / total_all_bills * 100), 2)) if total_all_bills > 0 else 0,
                        'rate_per_month': float(MAINTENANCE_RATE),
                    },
                    'security': {
                        'amount': security,
                        'percentage_of_categorized': float(round(security_pct, 2)),
                        'percentage_of_total': float(round((security / total_all_bills * 100), 2)) if total_all_bills > 0 else 0,
                        'rate_per_month': float(SECURITY_RATE),
                    },
                    'amenities': {
                        'amount': amenities,
                        'percentage_of_categorized': float(round(amenities_pct, 2)),
                        'percentage_of_total': float(round((amenities / total_all_bills * 100), 2)) if total_all_bills > 0 else 0,
                        'rate_per_month': float(AMENITIES_RATE),
                    },
                },
                'other_expenses': {
                    'amount': other_expenses,
                    'percentage_of_total': float(round((other_expenses / total_all_bills * 100), 2)) if total_all_bills > 0 else 0,
                },
                'service_statistics': service_stats,
                'payment_statistics': {
                    'total_bills': all_bills_aggregates['total_bills'],
                    'paid_bills': all_bills_aggregates['paid_bills'],
                    'pending_bills': all_bills_aggregates['pending_bills'],
                    'payment_rate': round((all_bills_aggregates['paid_bills'] / all_bills_aggregates['total_bills'] * 100), 2) if all_bills_aggregates['total_bills'] > 0 else 0,
                    'collection_rate': round((all_bills_aggregates['total_paid'] / total_all_bills * 100), 2) if total_all_bills > 0 else 0,
                },
                'calculation_info': {
                    'months_covered': months_count,
                    'units_with_bills': len(unit_ids_from_bills),
                }
            }
        
        # If include_other is true, add other_expenses to totalExpense
        if include_other:
            data['totalExpense'] = total_all_bills
            data['calculation_note'] = 'totalExpense includes other_expenses'
        else:
            data['calculation_note'] = 'totalExpense is sum of maintenance, security, and amenities only'
        
        serializer = ExpenseReflectionSerializer(data)
        return Response(serializer.data)
    
    def generate_chart_data(self, bills_queryset, chart_type, building_name=None, year=None, month=None):
        """Generate different types of chart data based on parameters"""
        
        if chart_type == 'pie':
            # Calculate totals
            total_all_bills = bills_queryset.aggregate(
                total=Coalesce(Sum('amount_due'), Decimal('0'))
            )['total']
            
            if total_all_bills == 0:
                return {
                    'type': 'pie',
                    'data': {
                        'labels': ['No Data'],
                        'datasets': [{
                            'data': [100],
                            'backgroundColor': ['#CCCCCC']
                        }]
                    },
                    'empty': True
                }
            
            # Get distinct units
            unit_ids_from_bills = bills_queryset.values_list('unit_id', flat=True).distinct()
            
            # Get assigned units
            assigned_units = AssignedUnit.objects.filter(
                unit_id__in=unit_ids_from_bills,
                deleted_at__isnull=True
            )
            
            # Fixed rates
            MAINTENANCE_RATE = Decimal('1500.00')
            SECURITY_RATE = Decimal('2000.00')
            AMENITIES_RATE = Decimal('2500.00')
            
            # Calculate expenses
            maintenance = Decimal('0.00')
            security = Decimal('0.00')
            amenities = Decimal('0.00')
            
            # If no assigned units found but building is specified, calculate based on all units
            if not assigned_units.exists() and building_name:
                units_in_building = Unit.objects.filter(building=building_name, deleted_at__isnull=True)
                for unit in units_in_building:
                    unit_bills = bills_queryset.filter(unit_id=unit.id)
                    if unit_bills.exists():
                        unit_months = unit_bills.annotate(
                            month=ExtractMonth('due_date'),
                            year=ExtractYear('due_date')
                        ).values('month', 'year').distinct().count()
                        
                        maintenance += MAINTENANCE_RATE * unit_months
                        security += SECURITY_RATE * unit_months
                        amenities += AMENITIES_RATE * unit_months
            else:
                for assigned_unit in assigned_units:
                    unit_bills = bills_queryset.filter(unit_id=assigned_unit.unit_id)
                    if unit_bills.exists():
                        unit_months = unit_bills.annotate(
                            month=ExtractMonth('due_date'),
                            year=ExtractYear('due_date')
                        ).values('month', 'year').distinct().count()
                        
                        if assigned_unit.maintenance:
                            maintenance += MAINTENANCE_RATE * unit_months
                        if assigned_unit.security:
                            security += SECURITY_RATE * unit_months
                        if assigned_unit.amenities:
                            amenities += AMENITIES_RATE * unit_months
            
            total_categorized = maintenance + security + amenities
            other_expenses = total_all_bills - total_categorized
            if other_expenses < Decimal('0'):
                other_expenses = Decimal('0')
            
            return {
                'type': 'pie',
                'data': {
                    'labels': ['Maintenance', 'Security', 'Amenities', 'Other Expenses'],
                    'datasets': [{
                        'data': [
                            float(round(maintenance, 2)),
                            float(round(security, 2)),
                            float(round(amenities, 2)),
                            float(round(other_expenses, 2))
                        ],
                        'backgroundColor': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0'],
                        'hoverBackgroundColor': ['#FF6384', '#36A2EB', '#FFCE56', '#4BC0C0']
                    }]
                },
                'total': float(total_all_bills)
            }

class YearlyExpenseAPIView(APIView):
    """
    API endpoint to get expense data grouped by year
    """
    
    def get(self, request, format=None):
        building_name = request.query_params.get('building')
        
        # Base queryset
        bills_queryset = MonthlyBill.objects.all()
        
        # Apply building filter if specified
        if building_name:
            units_in_building = Unit.objects.filter(building=building_name, deleted_at__isnull=True)
            unit_ids = units_in_building.values_list('id', flat=True)
            bills_queryset = bills_queryset.filter(unit_id__in=unit_ids)
        
        # Get distinct years
        years_data = bills_queryset.annotate(
            year=ExtractYear('due_date')
        ).values('year').distinct().order_by('-year')
        
        yearly_expenses = []
        
        for year_item in years_data:
            year = year_item['year']
            
            # Filter bills for this year
            yearly_bills = bills_queryset.filter(due_date__year=year)
            
            # Calculate totals for this year
            aggregates = yearly_bills.aggregate(
                total_all_bills=Coalesce(Sum('amount_due'), Decimal('0')),
                total_paid=Coalesce(
                    Sum('amount_due', filter=Q(payment_status=MonthlyBill.PaymentStatus.PAID)),
                    Decimal('0')
                ),
                total_unpaid=Coalesce(
                    Sum('amount_due', filter=Q(payment_status=MonthlyBill.PaymentStatus.PENDING)),
                    Decimal('0')
                )
            )
            
            total_all_bills = aggregates['total_all_bills']
            
            # Calculate categorized expenses
            CATEGORIZED_PERCENTAGE = Decimal('0.75')
            total_categorized = total_all_bills * CATEGORIZED_PERCENTAGE
            
            CATEGORY_DISTRIBUTION = {
                'maintenance': Decimal('0.40'),
                'security': Decimal('0.333'),
                'amenities': Decimal('0.267'),
            }
            
            maintenance = total_categorized * CATEGORY_DISTRIBUTION['maintenance']
            security = total_categorized * CATEGORY_DISTRIBUTION['security']
            amenities = total_categorized * CATEGORY_DISTRIBUTION['amenities']
            other_expenses = total_all_bills - (maintenance + security + amenities)
            
            yearly_expenses.append({
                'year': year,
                'maintenance': float(round(maintenance, 2)),
                'security': float(round(security, 2)),
                'amenities': float(round(amenities, 2)),
                'other_expenses': float(round(other_expenses, 2)),
                'totalExpense': float(round(maintenance + security + amenities, 2)),
                'total_all_bills': float(round(total_all_bills, 2)),
                'total_paid': float(round(aggregates['total_paid'], 2)),
                'total_unpaid': float(round(aggregates['total_unpaid'], 2)),
                'payment_rate': round((aggregates['total_paid'] / total_all_bills * 100), 2) if total_all_bills > 0 else 0,
            })
        
        return Response({
            'yearly_data': yearly_expenses,
            'building_filter': building_name if building_name else 'All Buildings',
            'total_years': len(yearly_expenses)
        })


class MonthlyExpenseAPIView(APIView):
    """
    API endpoint to get expense data for a specific year, grouped by month
    """
    
    def get(self, request, year, format=None):
        building_name = request.query_params.get('building')
        
        # Base queryset for the specific year
        bills_queryset = MonthlyBill.objects.filter(due_date__year=year)
        
        # Apply building filter if specified
        if building_name:
            units_in_building = Unit.objects.filter(building=building_name, deleted_at__isnull=True)
            unit_ids = units_in_building.values_list('id', flat=True)
            bills_queryset = bills_queryset.filter(unit_id__in=unit_ids)
        
        # Get monthly data
        monthly_data = bills_queryset.annotate(
            month=ExtractMonth('due_date')
        ).values('month').annotate(
            total_all_bills=Coalesce(Sum('amount_due'), Decimal('0')),
            total_paid=Coalesce(
                Sum('amount_due', filter=Q(payment_status=MonthlyBill.PaymentStatus.PAID)),
                Decimal('0')
            ),
            total_unpaid=Coalesce(
                Sum('amount_due', filter=Q(payment_status=MonthlyBill.PaymentStatus.PENDING)),
                Decimal('0')
            ),
            bill_count=Count('id')
        ).order_by('month')
        
        # Month names for display
        month_names = [
            'January', 'February', 'March', 'April', 'May', 'June',
            'July', 'August', 'September', 'October', 'November', 'December'
        ]
        
        monthly_expenses = []
        yearly_total = Decimal('0')
        yearly_paid = Decimal('0')
        yearly_unpaid = Decimal('0')
        
        for month_item in monthly_data:
            month_num = month_item['month']
            total_all_bills = month_item['total_all_bills']
            
            # Calculate categorized expenses
            CATEGORIZED_PERCENTAGE = Decimal('0.75')
            total_categorized = total_all_bills * CATEGORIZED_PERCENTAGE
            
            CATEGORY_DISTRIBUTION = {
                'maintenance': Decimal('0.40'),
                'security': Decimal('0.333'),
                'amenities': Decimal('0.267'),
            }
            
            maintenance = total_categorized * CATEGORY_DISTRIBUTION['maintenance']
            security = total_categorized * CATEGORY_DISTRIBUTION['security']
            amenities = total_categorized * CATEGORY_DISTRIBUTION['amenities']
            other_expenses = total_all_bills - (maintenance + security + amenities)
            
            monthly_expenses.append({
                'month_number': month_num,
                'month_name': month_names[month_num - 1] if 1 <= month_num <= 12 else f'Month {month_num}',
                'maintenance': float(round(maintenance, 2)),
                'security': float(round(security, 2)),
                'amenities': float(round(amenities, 2)),
                'other_expenses': float(round(other_expenses, 2)),
                'totalExpense': float(round(maintenance + security + amenities, 2)),
                'total_all_bills': float(round(total_all_bills, 2)),
                'total_paid': float(round(month_item['total_paid'], 2)),
                'total_unpaid': float(round(month_item['total_unpaid'], 2)),
                'bill_count': month_item['bill_count'],
                'payment_rate': round((month_item['total_paid'] / total_all_bills * 100), 2) if total_all_bills > 0 else 0,
            })
            
            yearly_total += total_all_bills
            yearly_paid += month_item['total_paid']
            yearly_unpaid += month_item['total_unpaid']
        
        # Calculate yearly categorized totals
        yearly_categorized = yearly_total * Decimal('0.75')
        yearly_maintenance = yearly_categorized * Decimal('0.40')
        yearly_security = yearly_categorized * Decimal('0.333')
        yearly_amenities = yearly_categorized * Decimal('0.267')
        yearly_other = yearly_total - (yearly_maintenance + yearly_security + yearly_amenities)
        
        return Response({
            'year': year,
            'monthly_data': monthly_expenses,
            'yearly_summary': {
                'maintenance': float(round(yearly_maintenance, 2)),
                'security': float(round(yearly_security, 2)),
                'amenities': float(round(yearly_amenities, 2)),
                'other_expenses': float(round(yearly_other, 2)),
                'totalExpense': float(round(yearly_maintenance + yearly_security + yearly_amenities, 2)),
                'total_all_bills': float(round(yearly_total, 2)),
                'total_paid': float(round(yearly_paid, 2)),
                'total_unpaid': float(round(yearly_unpaid, 2)),
                'payment_rate': round((yearly_paid / yearly_total * 100), 2) if yearly_total > 0 else 0,
            },
            'building_filter': building_name if building_name else 'All Buildings',
            'total_months': len(monthly_expenses)
        })
